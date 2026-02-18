# Fix for Python 3.8 compatibility with reportlab
import hashlib
import sys
if sys.version_info < (3, 9):
    # Monkey patch for Python 3.8 compatibility
    _original_md5 = hashlib.md5
    def _patched_md5(*args, **kwargs):
        kwargs.pop('usedforsecurity', None)
        return _original_md5(*args, **kwargs)
    hashlib.md5 = _patched_md5

import io
from datetime import datetime
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, PageBreak
)
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from .models import SystemConfig, InwardLot, ProcessProgram, CompanyDetail


def format_indian_number(number):
    """Format number with Indian comma notation (lakhs system)"""
    s = f"{float(number):.2f}"
    if '.' in s:
        integer_part, decimal_part = s.split('.')
    else:
        integer_part, decimal_part = s, "00"
    
    # Reverse the integer part for easier processing
    integer_part = integer_part[::-1]
    
    # Add commas: first at 3 digits, then every 2 digits
    parts = []
    if len(integer_part) > 3:
        parts.append(integer_part[:3])
        integer_part = integer_part[3:]
        while len(integer_part) > 2:
            parts.append(integer_part[:2])
            integer_part = integer_part[2:]
        if integer_part:
            parts.append(integer_part)
    else:
        parts.append(integer_part)
    
    # Reverse back and join
    formatted = ','.join(parts)[::-1]
    return f"{formatted}.{decimal_part}"


def generate_bill_pdf(bill):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )

    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        spaceAfter=20
    )

    heading_style = ParagraphStyle(
        'Heading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#333333'),
        spaceAfter=12,
        fontName='Helvetica-Bold'
    )

    # Get company details based on first program's quality type
    first_program = bill.programs.first()
    company_detail = None
    company_logo_path = None
    is_gstin_bill = False

    if first_program:
        lots = first_program.get_lots()
        if lots:
            quality_type = lots[0].quality_type
            is_gstin_bill = lots[0].is_gstin_registered
            try:
                company_detail = CompanyDetail.objects.get(quality_type=quality_type)
                if company_detail.logo:
                    company_logo_path = company_detail.logo.path
            except CompanyDetail.DoesNotExist:
                pass
    
    # Use company detail if available, otherwise fallback to SystemConfig
    if company_detail:
        company_name = company_detail.name
        company_address = company_detail.address
        company_phone = company_detail.phone
        company_email = company_detail.email
        company_gst = company_detail.gst_number or ""
    else:
        company_name = SystemConfig.get_config('COMPANY_NAME', 'ABC Textiles')
        company_address = SystemConfig.get_config('COMPANY_ADDRESS', 'Industrial Area, City')
        company_phone = SystemConfig.get_config('COMPANY_PHONE', '')
        company_email = SystemConfig.get_config('COMPANY_EMAIL', '')
        company_gst = SystemConfig.get_config('COMPANY_GST', '')
    
    # Add logo if available
    if company_logo_path:
        from reportlab.platypus import Image as RLImage
        try:
            logo = RLImage(company_logo_path, width=1.5*inch, height=1.5*inch, kind='proportional')
            elements.append(logo)
            elements.append(Spacer(1, 0.1*inch))
        except:
            pass  # If logo fails to load, continue without it
    
    elements.append(Paragraph(company_name, title_style))
    
    # Company contact info
    contact_info = f"{company_address}"
    if company_phone:
        contact_info += f" | Phone: {company_phone}"
    if company_email:
        contact_info += f" | Email: {company_email}"
    if is_gstin_bill and company_gst:  # Only show if registered
        contact_info += f" | GSTIN: {company_gst}"
    
    elements.append(Paragraph(contact_info, subtitle_style))
    elements.append(Spacer(1, 0.2*inch))

    bill_info_data = [
        ['Bill Number:', bill.bill_number, 'Date:', bill.bill_date.strftime('%d-%b-%Y')],
        ['Party Name:', bill.party.name, 'Total Programs:', str(bill.programs.count())],
    ]

    bill_info_table = Table(bill_info_data, colWidths=[1.5*inch, 2.5*inch, 1.5*inch, 1.5*inch])
    bill_info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#333333')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(bill_info_table)
    elements.append(Spacer(1, 0.3*inch))

    elements.append(Paragraph('Program Details', heading_style))

    table_data = [[
        'Program No', 'Design No', 'Quality Type', 'Lot Numbers',
        'Input (m)', 'Output (m)',
        'Rate (Rs./m)', 'Amount (Rs.)'
    ]]

    for program in bill.programs.all():
        lots = program.get_lots()
        lot_numbers = ', '.join([lot.lot_number for lot in lots])
        quality_type = lots[0].quality_type.name if lots else 'N/A'

        # Use centralized rate resolution
        rate = program.get_effective_rate()
        amount = program.output_meters * rate

        table_data.append([
            program.program_number,
            program.design_number,
            quality_type,
            lot_numbers,
            f"{program.input_meters:.2f}",
            f"{program.output_meters:.2f}",
            format_indian_number(rate),
            format_indian_number(amount)
        ])

    table = Table(table_data, colWidths=[
        1.1*inch, 0.9*inch, 1*inch, 1.3*inch,
        0.75*inch, 0.75*inch,
        0.85*inch, 1.1*inch
    ])

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a4a4a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f9f9f9')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 0.3*inch))

    # Conditional summary based on GSTIN status
    if is_gstin_bill:
        gst_rate = Decimal(SystemConfig.get_config('GST_RATE', '5.00'))
        summary_data = [
            ['Subtotal:', f"Rs. {format_indian_number(bill.subtotal)}"],
            [f'GST ({gst_rate}%):', f"Rs. {format_indian_number(bill.tax_total)}"],
            ['', ''],
            ['Grand Total:', f"Rs. {format_indian_number(bill.grand_total)}"],
        ]
        grand_total_row = 3
    else:
        summary_data = [
            ['Subtotal:', f"Rs. {format_indian_number(bill.subtotal)}"],
            ['', ''],
            ['Grand Total:', f"Rs. {format_indian_number(bill.grand_total)}"],
        ]
        grand_total_row = 2

    summary_table = Table(summary_data, colWidths=[5*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('FONTNAME', (0, grand_total_row), (-1, grand_total_row), 'Helvetica-Bold'),
        ('FONTSIZE', (0, grand_total_row), (-1, grand_total_row), 14),
        ('LINEABOVE', (0, grand_total_row), (-1, grand_total_row), 1.5, colors.black),
        ('TOPPADDING', (0, grand_total_row), (-1, grand_total_row), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))

    elements.append(summary_table)
    elements.append(Spacer(1, 0.5*inch))

    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )

    elements.append(Paragraph(
        f"Generated by Textile Inventory System on {datetime.now().strftime('%d-%b-%Y %I:%M %p')}",
        footer_style
    ))

    doc.build(elements)
    pdf_content = buffer.getvalue()
    buffer.close()
    return pdf_content


def generate_ledger_excel(party, start_date, end_date):
    from datetime import datetime as dt

    if isinstance(start_date, str):
        start_date = dt.strptime(start_date, '%Y-%m-%d').date()
    if isinstance(end_date, str):
        end_date = dt.strptime(end_date, '%Y-%m-%d').date()

    wb = Workbook()

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    ws1 = wb.active
    ws1.title = 'Inward Lots'

    ws1.append(['Party Ledger Report'])
    ws1.append([f'Party: {party.name}'])
    ws1.append([f'Period: {start_date.strftime("%d-%b-%Y")} to {end_date.strftime("%d-%b-%Y")}'])
    ws1.append([])

    ws1.merge_cells('A1:F1')
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A1'].alignment = center_align

    headers = ['Lot Number', 'Quality', 'Inward Date', 'Total Meters', 'Current Balance', 'Balance %']
    ws1.append(headers)

    header_row = ws1[5]
    for cell in header_row:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    lots = InwardLot.objects.filter(
        party=party,
        inward_date__range=[start_date, end_date]
    ).order_by('-inward_date')

    for lot in lots:
        balance_pct = lot.balance_percentage()
        ws1.append([
            lot.lot_number,
            lot.quality_type.name,
            lot.inward_date.strftime('%d-%b-%Y'),
            float(lot.total_meters),
            float(lot.current_balance),
            f"{balance_pct:.1f}%"
        ])

    for row in ws1.iter_rows(min_row=6, max_row=ws1.max_row):
        for cell in row:
            cell.border = border
            if cell.column in [4, 5]:
                cell.alignment = right_align

    # Auto-adjust column widths
    for idx, column in enumerate(ws1.columns, 1):
        max_length = 0
        column_letter = get_column_letter(idx)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws1.column_dimensions[column_letter].width = adjusted_width

    ws2 = wb.create_sheet(title='Programs')

    ws2.append(['Programs for Party'])
    ws2.append([f'Party: {party.name}'])
    ws2.append([f'Period: {start_date.strftime("%d-%b-%Y")} to {end_date.strftime("%d-%b-%Y")}'])
    ws2.append([])

    ws2.merge_cells('A1:I1')
    ws2['A1'].font = Font(bold=True, size=14)
    ws2['A1'].alignment = center_align

    headers = [
        'Program No', 'Design No', 'Quality Type', 'Lot Numbers',
        'Input (m)', 'Wastage (m)', 'Output (m)',
        'Status', 'Date'
    ]
    ws2.append(headers)

    header_row = ws2[5]
    for cell in header_row:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    programs = ProcessProgram.objects.filter(
        programlotallocation__lot__party=party,
        created_at__date__range=[start_date, end_date]
    ).distinct().order_by('-created_at')

    for program in programs:
        lots = program.get_lots()
        lot_numbers = ', '.join([lot.lot_number for lot in lots])
        quality_type = lots[0].quality_type.name if lots else 'N/A'

        ws2.append([
            program.program_number,
            program.design_number,
            quality_type,
            lot_numbers,
            float(program.input_meters),
            float(program.wastage_meters),
            float(program.output_meters),
            program.status,
            program.created_at.strftime('%d-%b-%Y')
        ])

    for row in ws2.iter_rows(min_row=6, max_row=ws2.max_row):
        for cell in row:
            cell.border = border
            if cell.column in [5, 6, 7]:  # Input, Wastage, Output columns
                cell.alignment = right_align
            elif cell.column in [3, 4]:  # Quality Type and Lot Numbers
                cell.alignment = left_align

    # Auto-adjust column widths
    for idx, column in enumerate(ws2.columns, 1):
        max_length = 0
        column_letter = get_column_letter(idx)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws2.column_dimensions[column_letter].width = adjusted_width

    ws3 = wb.create_sheet(title='Summary')

    ws3.append(['Summary Report'])
    ws3.append([f'Party: {party.name}'])
    ws3.append([f'Period: {start_date.strftime("%d-%b-%Y")} to {end_date.strftime("%d-%b-%Y")}'])
    ws3.append([])

    ws3.merge_cells('A1:B1')
    ws3['A1'].font = Font(bold=True, size=14)
    ws3['A1'].alignment = center_align

    total_inward = lots.aggregate(total=Sum('total_meters'))['total'] or Decimal('0')
    total_programs = programs.count()
    total_input = programs.aggregate(total=Sum('input_meters'))['total'] or Decimal('0')
    total_wastage = programs.aggregate(total=Sum('wastage_meters'))['total'] or Decimal('0')
    total_output = programs.aggregate(total=Sum('output_meters'))['total'] or Decimal('0')

    summary_data = [
        ['Total Inward Lots:', lots.count()],
        ['Total Inward Meters:', f"{float(total_inward):.2f}"],
        ['Total Programs:', total_programs],
        ['Total Input Meters:', f"{float(total_input):.2f}"],
        ['Total Wastage Meters:', f"{float(total_wastage):.2f}"],
        ['Total Output Meters:', f"{float(total_output):.2f}"],
        ['Wastage Percentage:', f"{(float(total_wastage)/float(total_input)*100):.2f}%" if total_input > 0 else "0%"],
    ]

    for row_data in summary_data:
        ws3.append(row_data)

    for row in ws3.iter_rows(min_row=5, max_row=ws3.max_row):
        row[0].font = Font(bold=True)
        row[0].alignment = Alignment(horizontal='left', vertical='center')
        row[1].alignment = right_align
        for cell in row:
            cell.border = border

    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    excel_content = buffer.getvalue()
    buffer.close()
    return excel_content


from django.db.models import Sum


def generate_comprehensive_ledger_excel(fiscal_year):
    """
    Generate a comprehensive ledger with separate sheets for each party.
    Includes: Inward Lots, Programs, Bills, and Summary for each party.

    Args:
        fiscal_year (int): The financial year (e.g., 2024 for FY 2024-25)

    Returns:
        bytes: Excel file content
    """
    from datetime import datetime as dt, date
    from .models import Party, InwardLot, ProcessProgram, Bill

    # Financial year runs from April 1st to March 31st
    start_date = date(fiscal_year, 4, 1)
    end_date = date(fiscal_year + 1, 3, 31)

    wb = Workbook()

    # Styling
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    subheader_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    subheader_font = Font(bold=True, size=10, color='000000')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    # Get all parties with activity in this fiscal year
    parties = Party.objects.filter(
        inwardlot__inward_date__range=[start_date, end_date]
    ).distinct().order_by('name')

    # Remove the default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Create overview sheet first
    overview_ws = wb.create_sheet(title='Overview', index=0)

    overview_ws.append(['Comprehensive Ledger Report'])
    overview_ws.append([f'Financial Year: FY {fiscal_year}-{str(fiscal_year+1)[-2:]}'])
    overview_ws.append([f'Period: {start_date.strftime("%d-%b-%Y")} to {end_date.strftime("%d-%b-%Y")}'])
    overview_ws.append([f'Generated on: {dt.now().strftime("%d-%b-%Y %H:%M")}'])
    overview_ws.append([])

    overview_ws.merge_cells('A1:D1')
    overview_ws['A1'].font = Font(bold=True, size=16)
    overview_ws['A1'].alignment = center_align

    overview_ws.append(['Party Name', 'Total Inward (m)', 'Total Programs', 'Total Billed Amount'])
    header_row = overview_ws[6]
    for cell in header_row:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # Generate sheet for each party
    for party in parties:
        # Get data for this party
        lots = InwardLot.objects.filter(
            party=party,
            inward_date__range=[start_date, end_date]
        ).order_by('-inward_date')

        programs = ProcessProgram.objects.filter(
            programlotallocation__lot__party=party,
            created_at__date__range=[start_date, end_date]
        ).distinct().order_by('-created_at')

        bills = Bill.objects.filter(
            programs__programlotallocation__lot__party=party,
            bill_date__range=[start_date, end_date]
        ).distinct().order_by('-bill_date')

        if not lots.exists():
            continue

        # Create safe sheet name (max 31 chars, no special chars)
        sheet_name = party.name[:28]
        sheet_name = ''.join(c if c.isalnum() or c in ' -_' else '_' for c in sheet_name)

        ws = wb.create_sheet(title=sheet_name)

        # Party Header
        ws.append([f'Party: {party.name}'])
        ws.append([f'FY {fiscal_year}-{str(fiscal_year+1)[-2:]} ({start_date.strftime("%d-%b-%Y")} to {end_date.strftime("%d-%b-%Y")})'])
        if party.contact:
            ws.append([f'Contact: {party.contact}'])
        if party.address:
            ws.append([f'Address: {party.address}'])
        ws.append([])

        ws.merge_cells(f'A1:F1')
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center_align

        current_row = ws.max_row + 1

        # Section 1: Inward Lots
        ws.append(['INWARD LOTS'])
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'].font = subheader_font
        ws[f'A{current_row}'].fill = subheader_fill
        ws[f'A{current_row}'].alignment = center_align
        ws[f'A{current_row}'].border = border

        headers = ['Lot Number', 'Quality', 'Inward Date', 'Total (m)', 'Balance (m)', 'Balance %']
        ws.append(headers)
        header_row = ws[ws.max_row]
        for cell in header_row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        for lot in lots:
            balance_pct = lot.balance_percentage()
            ws.append([
                lot.lot_number,
                lot.quality_type.name,
                lot.inward_date.strftime('%d-%b-%Y'),
                float(lot.total_meters),
                float(lot.current_balance),
                f"{balance_pct:.1f}%"
            ])

        for row in ws.iter_rows(min_row=current_row+2, max_row=ws.max_row):
            for idx, cell in enumerate(row):
                cell.border = border
                if idx >= 3:  # Numeric columns
                    cell.alignment = right_align

        ws.append([])
        current_row = ws.max_row + 1

        # Section 2: Programs
        ws.append(['PROGRAMS / CONVERSIONS'])
        ws.merge_cells(f'A{current_row}:I{current_row}')
        ws[f'A{current_row}'].font = subheader_font
        ws[f'A{current_row}'].fill = subheader_fill
        ws[f'A{current_row}'].alignment = center_align
        ws[f'A{current_row}'].border = border

        headers = ['Program No', 'Design No', 'Quality Type', 'Lot Numbers', 'Input (m)', 'Wastage (m)', 'Output (m)', 'Status', 'Date']
        ws.append(headers)
        header_row = ws[ws.max_row]
        for cell in header_row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        for program in programs:
            prog_lots = program.get_lots()
            lot_numbers = ', '.join([lot.lot_number for lot in prog_lots])
            quality_type = prog_lots[0].quality_type.name if prog_lots else 'N/A'

            ws.append([
                program.program_number,
                program.design_number,
                quality_type,
                lot_numbers,
                float(program.input_meters),
                float(program.wastage_meters),
                float(program.output_meters),
                program.status,
                program.created_at.strftime('%d-%b-%Y')
            ])

        for row in ws.iter_rows(min_row=current_row+2, max_row=ws.max_row):
            for idx, cell in enumerate(row):
                cell.border = border
                if idx in [4, 5, 6]:  # Numeric columns (Input, Wastage, Output)
                    cell.alignment = right_align
                elif idx in [2, 3]:  # Quality Type and Lot numbers
                    cell.alignment = left_align

        ws.append([])
        current_row = ws.max_row + 1

        # Section 3: Bills
        ws.append(['BILLING RECORDS'])
        ws.merge_cells(f'A{current_row}:G{current_row}')
        ws[f'A{current_row}'].font = subheader_font
        ws[f'A{current_row}'].fill = subheader_fill
        ws[f'A{current_row}'].alignment = center_align
        ws[f'A{current_row}'].border = border

        headers = ['Bill No', 'Bill Date', 'Programs', 'Subtotal', 'Tax', 'Grand Total', 'Payment Status']
        ws.append(headers)
        header_row = ws[ws.max_row]
        for cell in header_row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        for bill in bills:
            program_numbers = ', '.join([p.program_number for p in bill.programs.all()])

            ws.append([
                bill.bill_number,
                bill.bill_date.strftime('%d-%b-%Y'),
                program_numbers,
                float(bill.subtotal or 0),
                float(bill.tax_total or 0),
                float(bill.grand_total or 0),
                bill.payment_status
            ])

        for row in ws.iter_rows(min_row=current_row+2, max_row=ws.max_row):
            for idx, cell in enumerate(row):
                cell.border = border
                if idx in [3, 4, 5]:  # Numeric columns (Subtotal, Tax, Grand Total)
                    cell.alignment = right_align
                    cell.number_format = '₹#,##0.00'
                elif idx == 2:  # Program numbers
                    cell.alignment = left_align

        ws.append([])
        current_row = ws.max_row + 1

        # Section 4: Summary
        ws.append(['SUMMARY'])
        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws[f'A{current_row}'].font = subheader_font
        ws[f'A{current_row}'].fill = subheader_fill
        ws[f'A{current_row}'].alignment = center_align
        ws[f'A{current_row}'].border = border

        total_inward = lots.aggregate(total=Sum('total_meters'))['total'] or Decimal('0')
        total_input = programs.aggregate(total=Sum('input_meters'))['total'] or Decimal('0')
        total_wastage = programs.aggregate(total=Sum('wastage_meters'))['total'] or Decimal('0')
        total_output = programs.aggregate(total=Sum('output_meters'))['total'] or Decimal('0')
        total_billed = bills.aggregate(total=Sum('grand_total'))['total'] or Decimal('0')
        total_paid = bills.filter(payment_status='Paid').aggregate(total=Sum('grand_total'))['total'] or Decimal('0')
        total_outstanding = total_billed - total_paid

        summary_data = [
            ['Total Inward Lots:', lots.count()],
            ['Total Inward Meters:', f"{float(total_inward):.2f} m"],
            ['Total Programs:', programs.count()],
            ['Total Input Meters:', f"{float(total_input):.2f} m"],
            ['Total Wastage Meters:', f"{float(total_wastage):.2f} m"],
            ['Total Output Meters:', f"{float(total_output):.2f} m"],
            ['Wastage %:', f"{(float(total_wastage)/float(total_input)*100):.2f}%" if total_input > 0 else "0%"],
            [''],
            ['Total Bills:', bills.count()],
            ['Total Billed Amount:', f"₹{float(total_billed):,.2f}"],
            ['Total Paid:', f"₹{float(total_paid):,.2f}"],
            ['Outstanding Amount:', f"₹{float(total_outstanding):,.2f}"],
        ]

        for row_data in summary_data:
            ws.append(row_data)

        for row in ws.iter_rows(min_row=current_row+1, max_row=ws.max_row):
            if row[0].value:  # Skip empty rows
                row[0].font = Font(bold=True)
                row[0].alignment = left_align
                row[1].alignment = right_align
                for cell in row:
                    cell.border = border

        # Adjust column widths
        ws.column_dimensions['A'].width = 18  # Program No
        ws.column_dimensions['B'].width = 16  # Design No
        ws.column_dimensions['C'].width = 18  # Quality Type
        ws.column_dimensions['D'].width = 20  # Lot Numbers
        ws.column_dimensions['E'].width = 12  # Input
        ws.column_dimensions['F'].width = 12  # Wastage
        ws.column_dimensions['G'].width = 12  # Output
        ws.column_dimensions['H'].width = 12  # Status
        ws.column_dimensions['I'].width = 15  # Date

        # Add to overview sheet
        overview_ws.append([
            party.name,
            float(total_inward),
            programs.count(),
            float(total_billed)
        ])

    # Format overview sheet
    for row in overview_ws.iter_rows(min_row=7, max_row=overview_ws.max_row):
        for idx, cell in enumerate(row):
            cell.border = border
            if idx >= 1:  # Numeric columns
                cell.alignment = right_align
                if idx == 3:  # Amount column
                    cell.number_format = '₹#,##0.00'

    # Adjust overview column widths
    overview_ws.column_dimensions['A'].width = 30
    overview_ws.column_dimensions['B'].width = 18
    overview_ws.column_dimensions['C'].width = 18
    overview_ws.column_dimensions['D'].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    excel_content = buffer.getvalue()
    buffer.close()
    return excel_content
